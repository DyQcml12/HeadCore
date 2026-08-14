from __future__ import annotations

import csv
import hashlib
import json
import random
from collections import Counter
from pathlib import Path
from typing import Any, Iterable

from app.head.calibration import _fleiss_kappa


REVIEW_COLUMNS = (
    "reviewer_id",
    "item_id",
    "scenario_id",
    "user_input",
    "option_a",
    "option_b",
    "selected_option",
    "confidence",
    "notes",
)


def build_blind_review_package(
    scenario_results: list[dict[str, Any]],
    preferences: list[dict[str, Any]],
    *,
    seed: str = "hutao-headcore-blind-review-v1",
) -> dict[str, Any]:
    scenarios = {str(item["id"]): item for item in scenario_results}
    public_items: list[dict[str, str]] = []
    mappings: list[dict[str, str]] = []
    for preference in preferences:
        scenario_id = str(preference["scenario_id"])
        scenario = scenarios.get(scenario_id)
        if scenario is None:
            raise ValueError(f"blind review scenario missing: {scenario_id}")
        action_a = str(preference["preferred_action"])
        action_b = str(preference["rejected_action"])
        descriptions = {
            action: _candidate_objective(scenario, action) for action in (action_a, action_b)
        }
        ordered = [action_a, action_b]
        random.Random(f"{seed}:{preference['id']}").shuffle(ordered)
        item_id = _blind_item_id(seed, str(preference["id"]))
        public_items.append(
            {
                "item_id": item_id,
                "scenario_id": scenario_id,
                "user_input": str(scenario.get("user_input") or ""),
                "option_a": descriptions[ordered[0]],
                "option_b": descriptions[ordered[1]],
            }
        )
        planner_action = _planner_preferred_action(scenario, action_a, action_b)
        mappings.append(
            {
                "item_id": item_id,
                "preference_id": str(preference["id"]),
                "scenario_id": scenario_id,
                "option_a_action": ordered[0],
                "option_b_action": ordered[1],
                "reference_preferred_action": action_a,
                "headcore_preferred_action": planner_action,
            }
        )
    return {
        "schema_version": 1,
        "seed_fingerprint": hashlib.sha256(seed.encode()).hexdigest()[:12],
        "public_items": public_items,
        "mappings": mappings,
    }


def write_review_csv(package: dict[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=REVIEW_COLUMNS)
        writer.writeheader()
        for item in package["public_items"]:
            writer.writerow({**item, "reviewer_id": "", "selected_option": "", "confidence": "", "notes": ""})


def write_manifest(package: dict[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(package, ensure_ascii=False, indent=2), encoding="utf-8")


def load_review_rows(path: Path) -> list[dict[str, str]]:
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return [{key: str(value or "").strip() for key, value in row.items()} for row in csv.DictReader(handle)]
    if path.suffix.lower() == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - environment dependent
            raise ValueError("reading xlsx submissions requires openpyxl") from exc
        sheet = load_workbook(path, read_only=True, data_only=True)["Review Form"]
        values = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(values)]
        return [dict(zip(headers, [str(value or "").strip() for value in row])) for row in values]
    raise ValueError(f"unsupported review file: {path.suffix}")


def evaluate_blind_reviews(
    package: dict[str, Any], submissions: Iterable[tuple[Path, list[dict[str, str]]]]
) -> dict[str, Any]:
    mappings = {item["item_id"]: item for item in package["mappings"]}
    expected_ids = set(mappings)
    reviewers: dict[str, dict[str, str]] = {}
    confidence_values: list[int] = []
    notes_count = 0
    for path, rows in submissions:
        if not rows:
            raise ValueError(f"empty review submission: {path}")
        reviewer_ids = {row.get("reviewer_id", "").strip() for row in rows}
        if "" in reviewer_ids or len(reviewer_ids) != 1:
            raise ValueError(f"submission must contain one non-empty reviewer_id: {path}")
        reviewer_id = next(iter(reviewer_ids))
        if reviewer_id in reviewers:
            raise ValueError(f"duplicate reviewer submission: {reviewer_id}")
        item_ids = [row.get("item_id", "").strip() for row in rows]
        if len(item_ids) != len(set(item_ids)) or set(item_ids) != expected_ids:
            raise ValueError(f"submission item set is incomplete or duplicated: {path}")
        votes: dict[str, str] = {}
        for row in rows:
            choice = row.get("selected_option", "").strip().upper()
            if choice not in {"A", "B"}:
                raise ValueError(f"invalid selected_option for {row.get('item_id')}: {choice or 'blank'}")
            confidence = row.get("confidence", "").strip()
            if confidence:
                try:
                    confidence_value = int(confidence)
                except ValueError as exc:
                    raise ValueError(f"confidence must be an integer from 1 to 5: {confidence}") from exc
                if confidence_value not in range(1, 6):
                    raise ValueError(f"confidence must be an integer from 1 to 5: {confidence}")
                confidence_values.append(confidence_value)
            notes_count += int(bool(row.get("notes", "").strip()))
            votes[row["item_id"]] = choice
        reviewers[reviewer_id] = votes
    if len(reviewers) < 2:
        raise ValueError("blind review evaluation requires at least two reviewers")

    item_results: list[dict[str, Any]] = []
    vote_counts_for_kappa: list[dict[str, int]] = []
    unanimous = aligned = wins = ties = 0
    for item_id, mapping in mappings.items():
        choices = [votes[item_id] for votes in reviewers.values()]
        actions = [mapping[f"option_{choice.lower()}_action"] for choice in choices]
        counts = Counter(actions)
        vote_counts_for_kappa.append(dict(counts))
        ranking = counts.most_common()
        tied = len(ranking) > 1 and ranking[0][1] == ranking[1][1]
        consensus = None if tied else ranking[0][0]
        ties += int(tied)
        unanimous += int(len(counts) == 1)
        headcore_action = mapping["headcore_preferred_action"]
        is_aligned = consensus == headcore_action if consensus is not None else False
        aligned += int(is_aligned)
        wins += sum(action == headcore_action for action in actions)
        agreement = sum(count * (count - 1) for count in counts.values()) / (len(actions) * (len(actions) - 1))
        item_results.append(
            {
                "item_id": item_id,
                "scenario_id": mapping["scenario_id"],
                "vote_counts": dict(sorted(counts.items())),
                "consensus_action": consensus,
                "headcore_preferred_action": headcore_action,
                "headcore_aligned": is_aligned,
                "agreement": round(agreement, 4),
            }
        )
    item_count = len(item_results)
    judgment_count = item_count * len(reviewers)
    return {
        "reviewer_count": len(reviewers),
        "item_count": item_count,
        "judgment_count": judgment_count,
        "tie_count": ties,
        "raw_agreement": round(sum(item["agreement"] for item in item_results) / item_count, 4),
        "unanimous_rate": round(unanimous / item_count, 4),
        "fleiss_kappa": _fleiss_kappa(vote_counts_for_kappa),
        "headcore_consensus_alignment": round(aligned / (item_count - ties), 4) if item_count > ties else None,
        "headcore_vote_win_rate": round(wins / judgment_count, 4),
        "average_confidence": round(sum(confidence_values) / len(confidence_values), 2) if confidence_values else None,
        "notes_count": notes_count,
        "results": item_results,
    }


def _candidate_objective(scenario: dict[str, Any], action: str) -> str:
    candidates = [item for item in scenario.get("candidates", []) if item.get("action") == action]
    if not candidates:
        raise ValueError(f"candidate action missing for {scenario['id']}: {action}")
    return str(max(candidates, key=lambda item: float(item["score"]["total"]))["objective"])


def _planner_preferred_action(scenario: dict[str, Any], action_a: str, action_b: str) -> str:
    scores = {
        action: max(
            float(item["score"]["total"])
            for item in scenario["candidates"]
            if item.get("action") == action
        )
        for action in (action_a, action_b)
    }
    return max(scores, key=scores.get)  # type: ignore[arg-type]


def _blind_item_id(seed: str, preference_id: str) -> str:
    digest = hashlib.sha256(f"{seed}:{preference_id}".encode()).hexdigest()[:10].upper()
    return f"BR-{digest}"
